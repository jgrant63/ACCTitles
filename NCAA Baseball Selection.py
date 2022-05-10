# Welcome to the Universal Selection Indicator (USI), Jake's Proprietary NCAA Tournament Field Generation software.

# Import relevant packages
import pandas as pd
import sys
import numpy as np
import datetime as dt
import openpyxl
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import math
import requests

# Week of Rankings
week = 10

# Specify weights for input parameters. Values must sum to 1
w_adj_rpi = 0.525
w_wl = 0
w_non_con = 0.05
w_con = 0.05
w_road = 0
#w_last10 = 0.025
w_rpi25 = 0.125
w_rpi50 = 0.1
w_rpi100 = 0.075
w_rpi101 = 0.05
w_top100 = 0.025
w_bot150 = 0

weights = pd.Series([w_adj_rpi, w_wl, w_non_con, w_con, w_road, w_rpi25, w_rpi50, w_rpi100, w_rpi101, w_top100, w_bot150],index=['Adj. RPI Value','WL%','Non-Conf%','Conf%','Road WL%','RPI25%','RPI50%','RPI100%','RPI101+%','Top100%','Below150%'])
excel_path = '/Users/jakegrant/PycharmProjects/ACCTitles/NCAA_Baseball_Selection_Data.xlsx'

if round(weights.sum(),4) != 1.0000:
    out_str = "ERROR CODE 1: Weights equal %f. Enter values that equal 1." %weights.sum()
    print(out_str)
    sys.exit()
else:
    # Import data from NCAA Nitty Gritties website
    r = requests.get('https://stats.ncaa.org/selection_rankings/nitty_gritties/26625', headers={
        'Cookie': 'AKA_A2=A; X-Oracle-BMC-LBS-Route=e0e8f5ccc6c39fedaa6765ef3ac329941e557d93; _stats_session=BAh7B0kiD3Nlc3Npb25faWQGOgZFVEkiJTBmMWZiOGEyODhlZTEwNjgyNzdmNTAwOTk1OTVlZjEwBjsAVEkiEF9jc3JmX3Rva2VuBjsARkkiMTNzdVQrMVd3Sk16Mm04cGtBMkN5cE05QlVlb3ZVajhLL1d6TWcxcElFSWM9BjsARg%3D%3D--9481177a9fd3a9953a55baa1526f3e0e27c42ee5',
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/15.4 Safari/605.1.15'
    })
    scraped_data = pd.read_html(r.text, attrs={'id': 'selection_rankings_nitty_gritty_data_table'})
    scraped_week = scraped_data[0]

    # clean up the data by removing the bad columns from the HTML (ex: 04/28/2022 Result and Wins.1)
    valid_cols = []
    for col in scraped_week.columns:
        if ('Result' not in col) and ('Wins' not in col):
            valid_cols.append(col)

    scraped_vals = scraped_week.loc[:, valid_cols]

    book = load_workbook(excel_path)
    writer = pd.ExcelWriter(excel_path, engine='openpyxl')
    writer.book = book
    if f'Week_{week}' not in book.sheetnames:
        scraped_vals.to_excel(writer,sheet_name=f'Week_{week}')
    else:
        print('ERROR CODE 4: Sheet already exists. Code will resume.')
    writer.save()
    writer.close()

    # kick off the processing code
    d_season_jrpi = {}
    d_season_regs = {}
    last_in = {}
    for j in range(1,week+1):

        sheet = "Week_" + str(j)
        # Import data from NCAA Nitty Gritties website
        weekly_data = pd.read_excel(excel_path, sheet_name=sheet,engine='openpyxl')
        weekly_data = weekly_data.loc[:, ~weekly_data.columns.str.contains('^Unnamed')]

        # Calculate Win Percentage
        try:
            weekly_data['WL%'] = (weekly_data['WL'].str.split('-', expand=True)[0].astype(int) + .5*weekly_data['WL'].str.split('-', expand=True)[2].fillna(0).astype(int))/(weekly_data['WL'].str.split('-', expand=True)[0].astype(int) + weekly_data['WL'].str.split('-', expand=True)[1].astype(int) + weekly_data['WL'].str.split('-', expand=True)[2].fillna(0).astype(int))
        except KeyError:
            weekly_data['WL%'] = weekly_data['WL'].str.split('-', expand=True)[0].astype(int)/(weekly_data['WL'].str.split('-', expand=True)[0].astype(int) + weekly_data['WL'].str.split('-', expand=True)[1].astype(int))

        # Calculate Non-Conference Win Percentage
        try:
            weekly_data['Non-Conf%'] = (weekly_data['Non-Conf Record'].str.split('-', expand=True)[0].astype(int) + .5*weekly_data['Non-Conf Record'].str.split('-', expand=True)[2].fillna(0).astype(int))/(weekly_data['Non-Conf Record'].str.split('-', expand=True)[0].astype(int) + weekly_data['Non-Conf Record'].str.split('-', expand=True)[1].astype(int) + weekly_data['Non-Conf Record'].str.split('-', expand=True)[2].fillna(0).astype(int))
        except KeyError:
            weekly_data['Non-Conf%'] = weekly_data['Non-Conf Record'].str.split('-', expand=True)[0].astype(int) / (weekly_data['Non-Conf Record'].str.split('-', expand=True)[0].astype(int) + weekly_data['Non-Conf Record'].str.split('-', expand=True)[1].astype(int))

        # Calculate Conference Win Percentage
        try:
            weekly_data['Conf%'] = (weekly_data['Conf. Record'].str.split('-', expand=True)[0].astype(int) + .5*weekly_data['Conf. Record'].str.split('-', expand=True)[2].fillna(0).astype(int))/(weekly_data['Conf. Record'].str.split('-', expand=True)[0].astype(int) + weekly_data['Conf. Record'].str.split('-', expand=True)[1].astype(int) + weekly_data['Conf. Record'].str.split('-', expand=True)[2].fillna(0).astype(int))
        except KeyError:
            weekly_data['Conf%'] = weekly_data['Conf. Record'].str.split('-', expand=True)[0].astype(int) / (weekly_data['Conf. Record'].str.split('-', expand=True)[0].astype(int) + weekly_data['Conf. Record'].str.split('-', expand=True)[1].astype(int))

        # Calculate Road Win Percentage
        try:
            weekly_data['Road WL%'] = (weekly_data['Road WL'].str.split('-', expand=True)[0].astype(int) + .5*weekly_data['Road WL'].str.split('-', expand=True)[2].fillna(0).astype(int))/(weekly_data['Road WL'].str.split('-', expand=True)[0].astype(int) + weekly_data['Road WL'].str.split('-', expand=True)[1].astype(int) + weekly_data['Road WL'].str.split('-', expand=True)[2].fillna(0).astype(int))
        except KeyError:
            weekly_data['Road WL%'] = weekly_data['Road WL'].str.split('-', expand=True)[0].astype(int) / (weekly_data['Road WL'].str.split('-', expand=True)[0].astype(int) + weekly_data['Road WL'].str.split('-', expand=True)[1].astype(int))

        # Calculate Last 10 Games Win Percentage
        #try:
        #    weekly_data['Last 10%'] = (weekly_data['Last 10 Games'].str.split('-', expand=True)[0].astype(int) + .5*weekly_data['Last 10 Games'].str.split('-', expand=True)[2].fillna(0).astype(int))/(weekly_data['Last 10 Games'].str.split('-', expand=True)[0].astype(int) + weekly_data['Last 10 Games'].str.split('-', expand=True)[1].astype(int) + weekly_data['Last 10 Games'].str.split('-', expand=True)[2].fillna(0).astype(int))
        #except KeyError:
        #    weekly_data['Last 10%'] = weekly_data['Last 10 Games'].str.split('-', expand=True)[0].astype(int) / (weekly_data['Last 10 Games'].str.split('-', expand=True)[0].astype(int) + weekly_data['Last 10 Games'].str.split('-', expand=True)[1].astype(int))

        # Calculate Win Percentage vs. Top 25 RPI Teams
        try:
            weekly_data['RPI25%'] = (weekly_data['RPI 1-25'].str.split('-', expand=True)[0].astype(int) + .5*weekly_data['RPI 1-25'].str.split('-', expand=True)[2].fillna(0).astype(int))/(weekly_data['RPI 1-25'].str.split('-', expand=True)[0].astype(int) + weekly_data['RPI 1-25'].str.split('-', expand=True)[1].astype(int) + weekly_data['RPI 1-25'].str.split('-', expand=True)[2].fillna(0).astype(int))
        except KeyError:
            weekly_data['RPI25%'] = weekly_data['RPI 1-25'].str.split('-', expand=True)[0].astype(int) / (weekly_data['RPI 1-25'].str.split('-', expand=True)[0].astype(int) + weekly_data['RPI 1-25'].str.split('-', expand=True)[1].astype(int))

        # Calculate Win Percentage vs. 26-50 Ranked RPI Teams
        try:
            weekly_data['RPI50%'] = (weekly_data['RPI 26-50'].str.split('-', expand=True)[0].astype(int) + .5*weekly_data['RPI 26-50'].str.split('-', expand=True)[2].fillna(0).astype(int))/(weekly_data['RPI 26-50'].str.split('-', expand=True)[0].astype(int) + weekly_data['RPI 26-50'].str.split('-', expand=True)[1].astype(int) + weekly_data['RPI 26-50'].str.split('-', expand=True)[2].fillna(0).astype(int))
        except KeyError:
            weekly_data['RPI50%'] = weekly_data['RPI 26-50'].str.split('-', expand=True)[0].astype(int) / (weekly_data['RPI 26-50'].str.split('-', expand=True)[0].astype(int) + weekly_data['RPI 26-50'].str.split('-', expand=True)[1].astype(int))

        # Calculate Win Percentage vs. 51-100 Ranked RPI Teams
        try:
            weekly_data['RPI100%'] = (weekly_data['RPI 51-100'].str.split('-', expand=True)[0].astype(int) + .5*weekly_data['RPI 51-100'].str.split('-', expand=True)[2].fillna(0).astype(int))/(weekly_data['RPI 51-100'].str.split('-', expand=True)[0].astype(int) + weekly_data['RPI 51-100'].str.split('-', expand=True)[1].astype(int) + weekly_data['RPI 51-100'].str.split('-', expand=True)[2].fillna(0).astype(int))
        except KeyError:
            weekly_data['RPI100%'] = weekly_data['RPI 51-100'].str.split('-', expand=True)[0].astype(int) / (weekly_data['RPI 51-100'].str.split('-', expand=True)[0].astype(int) + weekly_data['RPI 51-100'].str.split('-', expand=True)[1].astype(int))

        # Calculate Win Percentage vs. >100 Ranked RPI Teams
        try:
            weekly_data['RPI101+%'] = (weekly_data['RPI 101+'].str.split('-', expand=True)[0].astype(int) + .5*weekly_data['RPI 101+'].str.split('-', expand=True)[2].fillna(0).astype(int))/(weekly_data['RPI 101+'].str.split('-', expand=True)[0].astype(int) + weekly_data['RPI 101+'].str.split('-', expand=True)[1].astype(int) + weekly_data['RPI 101+'].str.split('-', expand=True)[2].fillna(0).astype(int))
        except KeyError:
            weekly_data['RPI101+%'] = weekly_data['RPI 101+'].str.split('-', expand=True)[0].astype(int) / (weekly_data['RPI 101+'].str.split('-', expand=True)[0].astype(int) + weekly_data['RPI 101+'].str.split('-', expand=True)[1].astype(int))

        # Calculate Win Percentage vs. Top 100 RPI Teams
        try:
            weekly_data['Top100%'] = (weekly_data['vs TOP 100'].str.split('-', expand=True)[0].astype(int) + .5*weekly_data['vs TOP 100'].str.split('-', expand=True)[2].fillna(0).astype(int))/(weekly_data['vs TOP 100'].str.split('-', expand=True)[0].astype(int) + weekly_data['vs TOP 100'].str.split('-', expand=True)[1].astype(int) + weekly_data['vs TOP 100'].str.split('-', expand=True)[2].fillna(0).astype(int))
        except KeyError:
            weekly_data['Top100%'] = weekly_data['vs TOP 100'].str.split('-', expand=True)[0].astype(int) / (weekly_data['vs TOP 100'].str.split('-', expand=True)[0].astype(int) + weekly_data['vs TOP 100'].str.split('-', expand=True)[1].astype(int))

        # Calculate Win Percentage vs. Bottom 150 RPI Teams
        try:
            weekly_data['Below150%'] = (weekly_data['vs below 150'].str.split('-', expand=True)[0].astype(int) + .5*weekly_data['vs below 150'].str.split('-', expand=True)[2].fillna(0).astype(int))/(weekly_data['vs below 150'].str.split('-', expand=True)[0].astype(int) + weekly_data['vs below 150'].str.split('-', expand=True)[1].astype(int) + weekly_data['vs below 150'].str.split('-', expand=True)[2].fillna(0).astype(int))
        except KeyError:
            weekly_data['Below150%'] = weekly_data['vs below 150'].str.split('-', expand=True)[0].astype(int) / (weekly_data['vs below 150'].str.split('-', expand=True)[0].astype(int) + weekly_data['vs below 150'].str.split('-', expand=True)[1].astype(int))

        # clean up the data set
        valid_cols = []
        # print(weekly_data.columns)
        for col in weekly_data.columns:
            if ('Result' not in col) and ('Wins' not in col):
                valid_cols.append(col)

        weekly_vals = weekly_data.loc[:, valid_cols]
        # Calculate jRPI and Rank
        weekly_vals = weekly_vals.drop(['Team','Conference','SOS','Prev SOS','Prev Adj. RPI','Adj. RPI','RPI','RPI Value','Orig RPI Value','WL','Adj. Non-Conf RPI','Non-Conf Record','Conf RPI','Conf. Record','Road WL','Last 10 Games','RPI 1-25','RPI 26-50','RPI 51-100','RPI 101+','vs TOP 100','vs below 150','NC SOS'],axis=1,errors='ignore').fillna(0)
        weekly_jrpi = pd.DataFrame(weekly_vals.dot(weights), columns={'jRPI'})
        weekly_jrpi['Team'] = weekly_data['Team']
        weekly_jrpi['Conference'] = weekly_data['Conference']
        column_names = ['Team', 'Conference', 'jRPI']
        weekly_jrpi = weekly_jrpi.reindex(columns=column_names)
        weekly_jrpi['Rank'] = weekly_jrpi['jRPI'].rank(ascending=False)
        weekly_jrpi = weekly_jrpi.sort_values(by=['Rank']).reset_index(drop=True)

        # Find Autobids
        autos = weekly_jrpi.sort_values('Rank').groupby('Conference', as_index=False).first()['Team']
        weekly_jrpi['Autos'] = weekly_jrpi['Team'].isin(autos)

        # Find At Larges
        weekly_jrpi['At Large'] = weekly_jrpi.loc[weekly_jrpi['Autos'] == False, 'Rank']
        at_larges = weekly_jrpi.nsmallest(33, 'At Large', keep='first')['Team']
        last_in["last_{0}".format(j)] = at_larges.index[-1]
        weekly_jrpi['At Large'] = weekly_jrpi['Team'].isin(at_larges)

        # Determine Teams in Field
        weekly_jrpi['Field'] = weekly_jrpi['Autos'] | weekly_jrpi['At Large']
        tourney_field = weekly_jrpi.loc[weekly_jrpi['Field'] == True].reset_index(drop=True)

        # Regionals:
        team_pool = tourney_field
        d_regs = {}
        for i in range(16):
            seed_1 = team_pool.iloc[[0]]
            d_regs["reg_{0}".format(i)] = seed_1
            team_pool = team_pool[~team_pool['Team'].isin(seed_1['Team'])]

        for i in range(15,-1,-1):
            opt_1 = d_regs["reg_{0}".format(i)]
            valid_1 = opt_1['Conference']

            seed_2_ops = team_pool[~team_pool['Conference'].isin(valid_1)]
            seed_2 = seed_2_ops.iloc[[0]]
            d_regs["reg_{0}".format(i)] = pd.concat([opt_1,seed_2])
            team_pool = team_pool[~team_pool['Team'].isin(seed_2['Team'])]

        for i in range(16):
            opt_1_2 = d_regs["reg_{0}".format(i)].reset_index(drop=True)
            valid_1 = [opt_1_2['Conference'][0]]
            valid_2 = [opt_1_2['Conference'][1]]

            seed_3_ops = team_pool[~team_pool['Conference'].isin(valid_1)]
            seed_3_ops = seed_3_ops[~seed_3_ops['Conference'].isin(valid_2)]
            seed_3 = seed_3_ops.iloc[[0]]
            d_regs["reg_{0}".format(i)] = pd.concat([opt_1_2, seed_3])
            team_pool = team_pool[~team_pool['Team'].isin(seed_3['Team'])]

        for i in range(15, -1, -1):
            opt_1_2_3 = d_regs["reg_{0}".format(i)].reset_index(drop=True)
            valid_1 = [opt_1_2_3['Conference'][0]]
            valid_2 = [opt_1_2_3['Conference'][1]]
            valid_3 = [opt_1_2_3['Conference'][2]]

            seed_4_ops = team_pool[~team_pool['Conference'].isin(valid_1)]
            seed_4_ops = seed_4_ops[~seed_4_ops['Conference'].isin(valid_2)]
            seed_4_ops = seed_4_ops[~seed_4_ops['Conference'].isin(valid_3)]
            seed_4 = seed_4_ops.iloc[[0]]
            d_regs["reg_{0}".format(i)] = pd.concat([opt_1_2_3, seed_4]).reset_index(drop=True)
            team_pool = team_pool[~team_pool['Team'].isin(seed_4['Team'])]

        for key in d_regs:
            if len(d_regs[key]) < 4:
                out_str = "ERROR CODE 2: Invalid regional output. Adjust parameters for proper 4 team assignments (SEC error)."
                error_week = "Issue in Week %f" %j
                print(d_regs)
                print(out_str)
                print(error_week)
                sys.exit()

        d_season_jrpi["reg_{0}".format(j)] = weekly_jrpi.set_index('Team')
        d_season_regs["reg_{0}".format(j)] = d_regs

    # Assemble Master USI and Ranks Sheets
    season_jrpis = pd.DataFrame()
    season_ranks = pd.DataFrame()

    k = 0
    for key in d_season_jrpi:
        k += 1
        jrpi_rank = d_season_jrpi[key].drop(['Conference','Autos','At Large','Field'],axis=1)
        season_jrpis['Wk {}'.format(k)] = jrpi_rank['jRPI']
        season_ranks['Wk {}'.format(k)] = jrpi_rank['Rank']

    field_jrpis = season_jrpis[season_jrpis.index.isin(tourney_field['Team'])]
    field_ranks = season_ranks[season_ranks.index.isin(tourney_field['Team'])]

    # Plot of USI, Rank with Tech Highlighted
    # USI
    my_dpi = 96
    fig_val = 600
    fig1, ax = plt.subplots()
    plt.style.use('seaborn-darkgrid')
    plt.figure(figsize=(fig_val / my_dpi, fig_val / my_dpi), dpi=my_dpi)
    plt.plot(season_jrpis.T,marker='', color='grey', linewidth=1, alpha=0.4)
    plt.plot(season_jrpis.T['Georgia Tech'], marker='', color='#b3a369', linewidth=4, alpha=0.7)
    plt.ylim(bottom=0)
    plt.ylabel('USI Rating')
    plt.title('Weekly Change in Baseball USI Rating')
    plt.text(1, -0.1, 'Jake Grant, 2022', horizontalalignment='right',
             verticalalignment='center', transform=ax.transAxes)
    jRPI_fig = 'baseball_usi_plot.png'
    plt.savefig(jRPI_fig,dpi=my_dpi)
    #plt.show()

    # Rank
    fig2, ax = plt.subplots()
    plt.style.use('seaborn-darkgrid')
    plt.figure(figsize=(fig_val / my_dpi, fig_val / my_dpi), dpi=my_dpi)
    plt.plot(season_ranks.T, marker='', color='grey', linewidth=1, alpha=0.4)
    plt.plot(season_ranks.T['Georgia Tech'], marker='', color='#b3a369', linewidth=4, alpha=0.7)
    plt.plot(pd.DataFrame(list(last_in.items())).drop([0], axis=1), marker='', color='black', linewidth=2,
             alpha=0.7)
    plt.ylim(bottom=0)
    plt.gca().invert_yaxis()
    plt.ylabel('USI Ranking')
    plt.title('Weekly Change in Baseball USI Ranking')
    plt.text(1, -0.1, 'Jake Grant, 2022', horizontalalignment='right',
             verticalalignment='center', transform=ax.transAxes)
    rank_fig = 'baseball_rank_plot.png'
    plt.savefig(rank_fig,dpi=my_dpi)
    #plt.show()

    # Plot of USI, Rank for CURRENT Tournament Teams
    # USI
    my_dpi = 96
    fig_val = 600
    fig3, ax = plt.subplots()
    plt.style.use('seaborn-darkgrid')
    plt.figure(figsize=(fig_val / my_dpi, fig_val / my_dpi), dpi=my_dpi)
    plt.plot(field_jrpis.T, marker='', color='grey', linewidth=1, alpha=0.4)
    try:
        plt.plot(field_jrpis.T['Georgia Tech'], marker='', color='#b3a369', linewidth=4, alpha=0.7)
    except KeyError:
        print('ERROR CODE 3: Tech not in tournament field. Code will proceed.')
    plt.ylim(bottom=0)
    plt.ylabel('USI Rating')
    plt.title('Weekly Change in Baseball USI Rating for Tournament Teams')
    plt.text(1, -0.1, 'Jake Grant, 2022', horizontalalignment='right',
             verticalalignment='center', transform=ax.transAxes)
    jRPI_fig = 'baseball_usi_field_plot.png'
    plt.savefig(jRPI_fig, dpi=my_dpi)
    #plt.show()

    # Rank
    my_dpi = 96
    fig_val = 600
    fig4, ax = plt.subplots()
    plt.style.use('seaborn-darkgrid')
    plt.figure(figsize=(fig_val / my_dpi, fig_val / my_dpi), dpi=my_dpi)
    plt.plot(field_ranks.T, marker='', color='grey', linewidth=1, alpha=0.4)
    try:
        plt.plot(field_ranks.T['Georgia Tech'], marker='', color='#b3a369', linewidth=4, alpha=0.7)
        plt.plot(pd.DataFrame(list(last_in.items())).drop([0], axis=1), marker='', color='black', linewidth=2,
                 alpha=0.7)
    except KeyError:
        plt.plot(pd.DataFrame(list(last_in.items())).drop([0], axis=1), marker='', color='black', linewidth=2,
                 alpha=0.7)
    plt.ylim(bottom=0)
    plt.gca().invert_yaxis()
    plt.ylabel('USI Ranking')
    plt.title('Weekly Change in Baseball USI Ranking for Tournament Teams')
    plt.text(1, -0.1, 'Jake Grant, 2022', horizontalalignment='right',
             verticalalignment='center', transform=ax.transAxes)
    rank_fig = 'baseball_rank_field_plot.png'
    plt.savefig(rank_fig, dpi=my_dpi)
    #plt.show()

    # NEXT: ACC TEAM PLOT (WITH COLOR?)




    # Find Bubble Teams
    teams = weekly_jrpi['Team']
    teams = teams[~teams.isin(at_larges)]
    teams = teams[~teams.isin(autos)]

    # First Four Out
    first_four_out = teams[0:4].reset_index()
    first_four_out = first_four_out[['Team', 'index']]
    first_four_out['index'] = first_four_out['index'] + 1

    # Next Four Out
    next_four_out = teams[4:8].reset_index()
    next_four_out = next_four_out[['Team', 'index']]
    next_four_out['index'] = next_four_out['index'] + 1

    # Last Four In
    last_four_in = at_larges.iloc[-4:]
    last_four_in = last_four_in.iloc[::-1].reset_index()
    last_four_in = last_four_in[['Team', 'index']]
    last_four_in['index'] = last_four_in['index'] + 1

    # Next Four In
    next_four_in = at_larges.iloc[-8:-4]
    next_four_in = next_four_in.iloc[::-1].reset_index()
    next_four_in = next_four_in[['Team', 'index']]
    next_four_in['index'] = next_four_in['index'] + 1

    bubble = pd.concat([first_four_out, next_four_out, last_four_in, next_four_in],axis=1)
    bubble.columns = ['First Four Out', 'Rank', 'Next Four Out', 'Rank', 'Last Four In', 'Rank', 'Next Four In', 'Rank']

    # Writer Outputs to Excel
    tourney_projection = pd.concat(d_regs)
    out_xls_name = 'Week_' + str(week) + '_Baseball_Tourney_Prediction.xlsx'
    with pd.ExcelWriter(out_xls_name) as writer:
        tourney_projection.to_excel(writer, sheet_name='Projected Field')
        bubble.to_excel(writer,sheet_name='Bubble Teams')
        blank = pd.DataFrame()
        blank.to_excel(writer, sheet_name='USI Plot')
        blank.to_excel(writer, sheet_name='Rank Plot')

    # Print USI Plot to File
    wb = openpyxl.load_workbook(out_xls_name)
    ws = wb['USI Plot']
    img = openpyxl.drawing.image.Image(jRPI_fig)
    img.anchor = 'A1'  # Or whatever cell location you want to use.
    ws.add_image(img)
    wb.save(out_xls_name)

    # Print Rank Plot to File
    wb = openpyxl.load_workbook(out_xls_name)
    ws = wb['Rank Plot']
    img = openpyxl.drawing.image.Image(rank_fig)
    img.anchor = 'A1'  # Or whatever cell location you want to use.
    ws.add_image(img)
    wb.save(out_xls_name)

#### SCRATCH WORK ####


    # Unused Code to Add Plot Labels
    #num = 0
    #for i in df.values[9][1:]:
    #    num += 1
    #    name = list(df)[num]
    #    if name != 'y5':
    #        plt.text(10.2, i, name, horizontalalignment='left', size='small', color='grey')
    # Call Out Tech
    #plt.text(10.2, season_jrpis.T['Georgia Tech'].tail(1), 'Georgia Tech', horizontalalignment='left', size='small', color='#b3a369')

# Plot USI vs RPI, WL%, etc.
# plot of teams in vs. out (like show which ones have been in the whole team vs. not? idk
# ACC plots

